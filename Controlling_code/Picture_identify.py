import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import skimage
import sklearn
from skimage import filters
from skimage.io import imread, imshow
from skimage.color import rgb2gray, rgb2hsv
import cv2
import openpyxl
from scipy import interpolate
import scipy.signal as signal

font1 = {'family': 'Arial',
         'weight': 'normal',
         #"style": 'italic',
         'size': 14,
         }
class Point(object):
    def __init__(self,x,y):
        self.x = x
        self.y = y

    def getX(self):
        return self.x
    def getY(self):
        return self.y

def getGrayDiff(img,currentPoint,tmpPoint):
    return abs(int(img[currentPoint.x,currentPoint.y]) - int(img[tmpPoint.x,tmpPoint.y]))

def selectConnects(p):
    if p != 0:
        connects = [Point(-1, -1), Point(0, -1), Point(1, -1), Point(1, 0), Point(1, 1), \
                    Point(0, 1), Point(-1, 1), Point(-1, 0)]
    else:
        connects = [ Point(0, -1),  Point(1, 0),Point(0, 1), Point(-1, 0)]
    return connects

def regionGrow(img,seeds,p = 1):
    height, weight = img.shape
    seedMark = np.zeros(img.shape)
    seedList = []
    saveList = []
    for seed in seeds:
        seedList.append(seed)
    label = 1
    connects = selectConnects(p)
    while (len(seedList) > 0):
        currentPoint = seedList.pop(0)
        seedMark[currentPoint.x, currentPoint.y] = label
        for i in range(8):
            tmpX = currentPoint.x + connects[i].x
            tmpY = currentPoint.y + connects[i].y
            if tmpX < 0 or tmpY < 0 or tmpX >= height or tmpY >= weight:
                continue
            grayDiff = getGrayDiff(img, currentPoint, Point(tmpX, tmpY))
            if grayDiff != 0 and seedMark[tmpX, tmpY] == 0:
                seedMark[tmpX, tmpY] = label
                seedList.append(Point(tmpX, tmpY))
                saveList.append((tmpX, tmpY))
    return saveList







def Identify_TLC(image_name):
    #点的边缘
    Point_bound=[]

    #点的中心
    Point_center=[]

    #读取图像
    image = imread(image_name)
    #产生灰度图
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    for i in range(gray.shape[0]):
        for j in range(gray.shape[1]):
            if gray[i,j]<30:
                gray[i,j]=30
    #OSTU二值化
    ret1, th1 = cv2.threshold(gray, 0, 255, cv2.THRESH_OTSU)  #方法选择为THRESH_OTSU
    plt.subplot(111), plt.imshow(th1, "gray")
    plt.show()
    h=th1.shape[0] #600
    w=th1.shape[1] #800

    #找出边缘形状
    contours, hierarchy = cv2.findContours(th1, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

    new=np.zeros([h,w])
    #计算中心点
    x_plot=[]
    y_plot=[]
    Point_var=[]
    for i in range(len(contours)):
        x_total=[]
        y_total=[]
        if len(contours[i])>10:
            for j in range(np.squeeze(contours[i]).shape[0]):
                a=np.squeeze(contours[i])[j]
                new[a[1],a[0]]=255
                x_total.append(a[0])
                y_total.append(a[1])
            x_total=np.array(x_total)
            y_total=np.array(y_total)
            x_center=np.mean(x_total)
            y_center=np.mean(y_total)
            x_var=np.var(x_total)
            y_var=np.var(y_total)
            if 50<x_var+y_var<1000:
                x_plot.append(x_center)
                y_plot.append(y_center)
                Point_center.append([x_center,y_center])
                Point_var.append(x_var+y_var)

    print(sorted(Point_center))
    print(Point_var)
    # # #画出边缘
    plt.subplot(111), plt.imshow(new, "gray")
    plt.scatter(np.array(x_plot),np.array(y_plot),c='red')
    plt.draw()
    plt.pause(5)  # 间隔的秒数：6s
    plt.close()

    return Point_center

def Identify_TLC_Hight(picture_name,dx=125,dboundary=100,dinterval=60):
    image = imread(picture_name)[:,750:1750]
    #产生灰度图
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    # plt.imshow(gray,'gray')
    # plt.show()
    x_shade=np.zeros([gray.shape[1]])
    y_shade=np.zeros([gray.shape[0]])
    x_white=[]
    y_white=[]
    for i in range(gray.shape[1]):
        x_shade[i] = np.sum(gray[:,i])/gray.shape[0]
        if x_shade[i]>40:
            x_white.append(i)
    for i in range(gray.shape[0]):
        y_shade[i] = np.sum(gray[i]) / gray.shape[1]
        if y_shade[i] > 15:
            y_white.append(i)
    # print(x_white[0],x_white[-1])
    # print(y_white[0],y_white[-1])

    x_1_center = x_white[0] + dboundary
    x_2_center = x_1_center + dx
    x_3_center = x_2_center + dx
    x_4_center = x_3_center + dx
    y_limit = 1500
    y_1_shade = np.zeros(y_limit)
    y_2_shade = np.zeros(y_limit)
    y_3_shade = np.zeros(y_limit)
    y_4_shade = np.zeros(y_limit)
    x_1_shade = np.zeros([2 * dinterval])
    x_2_shade = np.zeros([2 * dinterval])
    x_3_shade = np.zeros([2 * dinterval])
    x_4_shade = np.zeros([2 * dinterval])
    for i in range(y_limit):
        y_1_shade[i] = np.sum(gray[i, x_1_center - dinterval:x_1_center + dinterval]) / (2 * dinterval)
        y_2_shade[i] = np.sum(gray[i, x_2_center - dinterval:x_2_center + dinterval]) / (2 * dinterval)
        y_3_shade[i] = np.sum(gray[i, x_3_center - dinterval:x_3_center + dinterval]) / (2 * dinterval)
        y_4_shade[i] = np.sum(gray[i, x_4_center - dinterval:x_4_center + dinterval]) / (2 * dinterval)


    for i in range(2*dinterval):
        x_1_shade[i]=np.sum(gray[:,x_1_center-dinterval+i])/(gray.shape[0])
        x_2_shade[i]=np.sum(gray[:,x_2_center-dinterval+i])/(gray.shape[0])
        x_3_shade[i]=np.sum(gray[:,x_3_center-dinterval+i])/(gray.shape[0])
        x_4_shade[i]=np.sum(gray[:,x_4_center-dinterval+i])/(gray.shape[0])

    plt.figure(1,figsize=(3,4),dpi=300)
    plt.style.use('ggplot')
    plt.plot(y_1_shade,c='black')
    plt.yticks(fontproperties='Arial', size=9)
    plt.xticks(fontproperties='Arial', size=9)
    plt.savefig("TLC_hight.pdf",dpi=300,bbox_inches='tight')
    plt.show()

    y_hight=np.array([np.where(y_1_shade==np.min(y_1_shade))[0][0],np.where(y_2_shade==np.min(y_2_shade))[0][0],
             np.where(y_3_shade==np.min(y_3_shade))[0][0],np.where(y_4_shade==np.min(y_4_shade))[0][0]])
    x_hight = np.array([x_1_center,
                        x_2_center,
                        x_3_center,
                        x_4_center])

    true_hight=(y_white[-1]-y_hight)*25.1/604-8.1
    #画图
    plt.subplot(111), plt.imshow(image)
    plt.scatter(x_hight,y_hight,c='red')
    plt.axis('off')
    plt.savefig('biaozhu.pdf',dpi=300)
    return true_hight

def Identify_Eluent_Hight(picture_name):
    #print(Identify_TLC_true(r"C:\Users\xxhhss6910\PycharmProjects\AUBO\Save_Fig\Experiment_8171117\board_2.jpg"))

    image = imread(picture_name)
    #产生灰度图
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    y_shade = np.zeros([gray.shape[0]])
    x_shade=np.zeros([gray.shape[1]])
    x_white=[]
    for i in range(gray.shape[1]):
        x_shade[i] = np.sum(gray[:,i])/gray.shape[0]
        if x_shade[i]>40:
            x_white.append(i)
    for i in range(gray.shape[0]):
        y_shade[i] = np.sum(gray[i, x_white[0]+200:x_white[-1]-200]) / (x_white[-1]-x_white[0]-400)

    plt.figure(1, figsize=(3, 4), dpi=300)
    plt.style.use('ggplot')
    plt.plot(y_shade,c='black')
    plt.yticks(fontproperties='Arial', size=9)
    plt.xticks(fontproperties='Arial', size=9)
    plt.savefig("guang.pdf",dpi=300,bbox_inches='tight')
    plt.show()
    y_high=[]
    for i in range(len(y_shade)):
        if y_shade[i] > max(y_shade)-4:
            y_high.append(i)
    # plt.plot(y_shade)
    # plt.show()
    #y_hight=np.where(y_shade==np.max(y_shade))[0][0]-55  #上边缘
    y_hight=y_high[-1]
    #print(y_hight)
    hight=(1944-y_hight)*0.0267+17.82-8.1
    return hight

def Calculate_RF(potential_hight,eluent_hight):
    RF=[]
    tlc_hight=potential_hight
    for i in range(len(tlc_hight)):
        if tlc_hight[i]/eluent_hight<0:
            RF.append(0)
        if tlc_hight[i]/eluent_hight>1:
            RF.append(1)
        else:
            RF.append(tlc_hight[i]/eluent_hight)
    return RF

import time
a=time.time()
potential_hight=Identify_TLC_Hight(r'D:\pycharm project\Picture_identify\TLC_identify\board_3_1.jpg')
eluent_hight=Identify_Eluent_Hight(r'D:\pycharm project\Picture_identify\TLC_identify\line_3_1.jpg')
RF=Calculate_RF(potential_hight,eluent_hight)
print(RF)
b=time.time()
print(b-a)
print(potential_hight)
print(eluent_hight)
# Dir_name='Experiment_9171338/'
# COL=7
# NUM=2
# for i in range(1,7):
#     for j in range(1,4):
#         TLC_name=Dir_name+'board_%d_%d.jpg'%(i,j)
#         Eluent_name=Dir_name+"line_%d_%d.jpg"%(i,j)
#         potential_hight=Identify_TLC_Hight(TLC_name)
#         eluent_hight=Identify_Eluent_Hight(Eluent_name)
#         RF=Calculate_RF(potential_hight,eluent_hight)

        # x1 = openpyxl.load_workbook('TLC_data.xlsx')  # 找到需要xlsx文件的位置
        #
        # sheet_name = x1.get_sheet_names()
        # sheet_info = x1.get_sheet_by_name(sheet_name[0])
        # sheet_info.cell(1,COL).value = Dir_name
        # for k in range(4):
        #     sheet_info.cell(NUM,COL).value = RF[k]
        #     NUM+=1
        # x1.save('TLC_data.xlsx')
        # print(RF)




#523,465==4.92mm   1pix=0.0848mm
#600  6.5296mm 相片离底部2.10mm
#(600-x)*0.0848+2.10mm